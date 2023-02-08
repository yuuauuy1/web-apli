#必要ライブラリのインポート
import openpyxl 
import csv
from fractions import Fraction  #有理数(分数で表現できるような値)を扱うためのモジュール

#--------------------ファイルを開く-----------------------------------
ing_wb = openpyxl.load_workbook("recipe_ingredients.xlsx") #レシピの材料と量の一覧表
ing_ws = ing_wb.worksheets[0] #ワークシート１枚目を選択

index_wb = openpyxl.load_workbook("recipe_index.xlsx") #レシピ名、ＵＲＬ、材料、人数分の表一覧
index_ws = index_wb.worksheets[0]

#コマンドラインに表示する言葉のリストをつくる
with open("word_list.csv" , "r", encoding = "shift_jis") as f:
    reader = csv.reader(f)
    word_list = []
    for row in reader:
        row[0].replace("\u3000", " ")
        word_list.append(row[0])
  
#index.レシピ名.URL.人数.レシピ名と材料が一緒になった列のリストを作っておく
index_name_lists=[] 
for row in index_ws.iter_rows(min_row =2, min_col=1, max_col=6):
    index_name =[]
    index_name.append(row[0].value) #index列
    index_name.append(row[1].value) #レシピ名
    index_name.append(row[2].value) #URL
    index_name.append(row[3].value) #人数
    index_name.append(row[5].value) #レシピ名と材料が一緒になった列　

    index_name_lists.append(index_name)

#index.材料.量が一緒になったリストを作っておく
index_ing_lists=[]
for row in ing_ws.iter_rows(min_row =2, min_col=1, max_col=3):
    index_ing =[]
    index_ing.append(row[0].value) #index列
    index_ing.append(row[1].value) #材料
    index_ing.append(row[2].value) #量

    index_ing_lists.append(index_ing)

# 検索文字が正しく入力されているか判定する関数
def validate_word():  #検討ポイント
    word = str(input(word_list[0]))
    word_result =  word.isalpha() #ひらがな／カタカナ／漢字が文字判定(True)される。数字は除外(False)
    if word_result :
        search(word)
    else:
        print(word_list[1]) #日本語以外の文字が入力されました      
        validate_word()  

def search(word): # 検索文字が一致するセルの座標をrecipe_indexファイルから検索   
    title_list = [] #検索件数を把握するためにtitle_list にtitle格納する
    for index_name in index_name_lists: 
         # 検索ワードと一致した時        
        if word in index_name[4]:
          
            # index と URLを取得する
            index = index_name[0]
            title = index_name[1]
            url = index_name[2]
            people_n = index_name[3]
            title_list.append(title)
            #print("+++++++++++++++++++++++++++++++++++++++++")
            #print("index："+ str(index))
            #print("レシピ名："+ title)
            #print(str(people_n)+ "　人分のレシピです")
            #print("URL："+ url)
            message()

            for index_ing in index_ing_lists:   #リストとして持っておく
                if str(index) in str(index_ing[0]):
                    ingredient = index_ing[1]
                    amount =  index_ing[2]           
                    print("{"+ str(ingredient)+"："+str(amount)+"}")

    print( str( len(title_list))+ "件のレシピが見つかりました。" )

    if len(title_list) == 0: # 検索結果無い場合、検索ワード入力に戻る
        print(word_list[2])#検索ワードの文字表記を変えると検索結果変わるかも・・・
        validate_word()
    
    else:                    # 検索結果がある場合、レシピのindex選択に進む
        number()

def message():
    print("+++++++++++++++++++++++++++++++++++++++++")
    print("index："+ str(index))
    print("レシピ名："+ title)
    print(str(people_n)+ "　人分のレシピです")
    print("URL："+ url)
    
def number(): # 買い物リストに入れるレシピの index が 正しい数字か判定
    recipe_index = input( word_list[3] ) #買い物リストに入れるレシピの　indexを入力してください
    index_result = recipe_index.isdigit() # 正の整数か判定
    
    if index_result:
        if 0 <= int(recipe_index) <= 2031:
            index_list.append(int(recipe_index))
            end()        
        else:
            print(word_list[4]) #indexにない数字が入力されました
            number()
    else:
        print(word_list[5]) #数字以外の文字が入力されました
        number()

def end(): # レシピ検索継続するか 終了するか入力
    input_n = input(word_list[6]) #まだレシピ検索しますか？Yes:1,No:0
    n_result = input_n.isdigit() # 正の整数か判定
    if n_result:
        
        if int(input_n) == 1:
            validate_word()
        
        elif int(input_n) == 0:
            print("レシピ選択を終わります。　現在" + str(len(index_list)) + "個のレシピが選択されています。")
        else:
            print(word_list[7]) #1 or 0で選択してください
            end()
    else:
        print(word_list[5]) #数字以外が入力されました
        end()

index_list =[] # number関数 で使っている レシピindexのリスト
validate_word()

print("選択されたレシピ index : " + str(index_list))
print(word_list[8]) #それでは買い物リストを作りましょう

#--------------選択したレシピの材料と量をそれぞれ辞書にする----------------
def people(): # ユーザーが作りたい人数が整数であるか判定し、欲しい人数の量に変換する関数
        human_n = input(word_list[9])  #ユーザーが作りたい人数を入力
        human_n_result = human_n.isdigit()
        if human_n_result: #追加121-136インデント
            ingredients_a =[]
            amount_a =[]   

            for row in ing_ws.iter_rows(min_col=1, max_col=6):

                if int(index) == row[0].value:
                    ingredients_a.append(row[5].value)
                    num = Fraction(row[3].value)
                    num = num * Fraction(int(human_n), int(people_n)) #欲しい人数の量に変換
                    amount_a.append(num)
                    dic_a= dict(zip(ingredients_a, amount_a)) 
            dic_list.append(dic_a) # 辞書同士を連結させるために、辞書のリストを作る
        else:
            print(word_list[5])
            people()

dic_list = []
count = 1
for index in index_list:
    
    #レシピ一覧から　Indexと　何人分のレシピか取得する
    for row in index_ws.iter_rows(min_col=1, max_col=4):

        if int(index) == row[0].value : 
           people_n = row[3].value #何人分のレシピか
           title = row[1].value #レシピの名前
           print("*********************************")
           print(str(count) + "番目, レシピNo：" +str(index)+ " ,　こちらは " +str(people_n) + " 人分のレシピです")
           print("レシピ名　「　"+ title+"　」")
    
    #材料と量の一覧から、材料と欲しい人数分の量に直したものを取得
    print(word_list[10])#何人分の材料を買いたいですか？
    people()
    
    #print(dic_a)
    count += 1
      
#print(dic_list)

#-------------------材料と量一覧を取得する----------------------------------
new_dic = {}
for dic in dic_list:              

    for key in dic.keys() : 

        if key in new_dic: # dicとnew_dicに同じkeyがある場合はvalueを足す
            num_1 = Fraction(new_dic.get(key) or 0) #分数の分量があるのでFraction型に変換する      
            num_2 = Fraction(dic.get(key) or 0) #分量Noneがある可能性あるので　「or 0」 を入力      
            sum = num_1 + num_2          
            new_dic[key] = sum
        else: # 同じkeyがない場合は　dicのvalueをそのまま採用
            new_dic[key] = dic[key]

#-----------------お買い物リスト表示---------------------------------------
print(word_list[11]) #----買い物リスト---
print(word_list[12])#調味料は０と表示される
print(word_list[13])#（）の中は単位です

for key, sum in new_dic.items():
    print(key + " : " + str(sum))
  